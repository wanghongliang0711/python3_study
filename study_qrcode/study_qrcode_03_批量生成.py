"""
@author: blake.wang
@file: study_qrcode_03_批量生成.py
@time: 2022/11/9 13:57 
"""
import os
import time

import qrcode
from openpyxl import load_workbook


# 创建文件夹
def make_dir(path):
    try:
        if not os.path.isdir(path):
            os.makedirs(path)
    except Exception as e:
        print("make_dir error . message: ", e)


def get_excel_data(path):
    result = []
    try:
        wb = load_workbook(path)
        sheet1 = wb[wb.sheetnames[0]]
        sheet1_rows = sheet1.max_row
        if sheet1_rows == 1:
            print("Excel has no data !!!")
        else:
            for i in range(1, sheet1_rows + 1):
                cell_str = sheet1.cell(i, 1).value
                cell_str = cell_str if cell_str is not None else ""
                if cell_str != "":
                    result.append(cell_str)
        return result
    except Exception as e:
        print("get_excel_data error . message: ", e)
        return result


def main():
    try:
        while True:
            input1 = str(input(r'Please enter a Excel(.xlsx) file path:'))
            user_input1 = input1.replace('"', '').replace("'", '')
            if os.path.exists(user_input1) and (user_input1.endswith(".xlsx") or user_input1.endswith(".XLSX")):
                data = get_excel_data(user_input1)
                folder_name = time.strftime("qrcode_%Y%m%d_%H%M%S", time.localtime())
                folder_path = os.path.join(os.getcwd(), folder_name)
                make_dir(folder_path)
                for one in data:
                    one = str(one)
                    img = qrcode.make(one)
                    img_path = os.path.join(folder_path, one+'.png')
                    img.save(img_path)
                print("结果地址： ", folder_path)
                print("\n")
            else:
                print("excel 路径错误，或者excel不是 .xlsx 结尾！！！")
    except Exception as e:
        print("Run error . message: ", e)


if __name__ == '__main__':
    main()
