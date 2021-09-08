#-*-coding:utf-8 -*-
'''
Created on 2019年5月15日

@author: 14668
'''
import datetime
import os

from openpyxl import Workbook  # 将数据写入Excel
from openpyxl import load_workbook  # 读取Excel文件


#从1开始，不是0
#读取Excel文件    
# 默认可读写，若有需要可以指定write_only和read_only为True
filepath = os.getcwd()+'\\demo.xlsx'
wb = load_workbook(filepath)

#获得所有sheet的名称
print(wb.get_sheet_names())

# 根据sheet名字获得sheet
# a_sheet = wb.get_sheet_by_name('MessageErrorHandle')
a_sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
# 获得sheet名
print(a_sheet.title)

# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = wb.active 
print(sheet)#<Worksheet "MessageErrorHandle">

#获取单元格
# 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
b4 = sheet['B1']
print(b4.value)
print(b4.column)
print(b4.row)

# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
b4_too = sheet.cell(row=1, column=2)
print(b4_too.value)

# 获得最大列和最大行
print(sheet.max_row)
print(sheet.max_column)


#获取行和列
# 因为按行，所以返回A1, B1, C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value)

#因为sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引，
for cell in list(sheet.rows)[1]:
    print('**')
    print(cell.value)

print('*************************************************')

#############将数据写入Excel############################

#若要指定只写模式，可以指定参数write_only=True。一般默认的可写可读模式就可以了。
wkbk = Workbook() 

print(wkbk.get_sheet_names())# # 提供一个默认名叫Sheet的表，office2016下新建提供默认Sheet1

#默认表sheet1
ws = wkbk.active
# 直接赋值就可以改工作表的名称
ws.title = 'my name is Sheet1'
#创建sheet2表
ws1 = wkbk.create_sheet('sheet2')

# 新建一个工作表，可以指定索引，适当安排其在工作簿中的位置
# wkbk.create_sheet('Data', index=1) # 被安排到第二个工作表，index=0就是第一个位置

# 删除某个工作表
# wb.remove(sheet)
# del wb[sheet]

#写入单元格
dttm = datetime.datetime.strptime("140623 135630.508","%y%m%d %H%M%S.%f")
dttm2 = datetime.datetime.strptime("135630","%H%M%S")
print(dttm)
print(dttm2)
# ws['A1'] = dttm
cell = ws['A1']
cell.value = dttm
cell.number_format = "MM-DD HH:MM:SS.000"

ws.cell(3,3).value = "140623 135630.508"
ws.cell(3,3).number_format = "MM-DD HH:MM:SS.000"


ws['A2'] = dttm2
ws['A2'] = 6699
# cell = ws['A']
# cell.number_format = "MM-DD HH:MM:SS.000"
'''
# B9处写入平均值
sheet['B9'] = '=AVERAGE(B2:B8)'
但是如果是读取的时候需要加上data_only=True这样读到B9返回的就是数字，如果不加这个参数，返回的将是公式本身'=AVERAGE(B2:B8)'
'''

#append函数
#可以一次添加多行数据，从第一行空白行开始（下面都是空白行）写入。
# 添加一行
row = [1 ,2, 3, 4, 5]
ws.append(row)
ws.append(['&&&&', b'freg'])
# 添加多行
rows = [
    ['Number', 'data1', 'data2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]

for data in rows:
    ws.append(data)
    
nrow = 15
value = [6,8,'tt']
for valu in value:
    ws.cell(nrow+1,1).value = valu
    ws.cell(row=nrow+1,column=2,value='sheet2表2行2列的值').value
    nrow = nrow + 1

for i in range(10):
    ws['H%d'%(i+1)].value = 'good'
    

#保存文件
#所有的操作结束后，一定记得保存文件。指定路径和文件名，后缀名为xlsx
path66 = os.getcwd()+'\\write_test.xlsx'
wkbk.save(path66)

# 以前从来没有 close 过，以后记得close
wkbk.close()
