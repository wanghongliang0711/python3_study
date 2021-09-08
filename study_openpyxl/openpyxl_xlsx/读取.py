"""
@author: wanghongliang
@file: 读取.py
@time: 2021/9/6 10:24 
"""
import os
from openpyxl import load_workbook  # 读取Excel文件


filepath = os.getcwd()+'\\write.xlsx'
wb = load_workbook(filepath)



print(wb.sheetnames)
a_sheet = wb[wb.sheetnames[0]]

print(a_sheet.title)

print(a_sheet.rows)
print(list(a_sheet.rows))


print("////////////")

for row in a_sheet.rows:
    print(row)

print("////////////")


sheet_names = wb.sheetnames
print("这个excel共有 %d 个sheet分别为 %s " % (len(sheet_names), sheet_names))

for k in range(0, len(sheet_names)):
    ws = wb[sheet_names[k]]
    # print(ws.title)
    rows = ws.max_row  # 获取这个sheet的行数最大值
    columns = ws.max_column # 获取这个sheet的列数最大值

    # print(rows)
    # print(columns)
    """
    只有A1 有数据 rows columns 都是 1
    一个数据没有 rows columns 也都是 1
    """
    if rows == 1 and columns == 1:
        print("break==========")
        break
    else:
        print("正在打印 %s sheet数据 共有 %d 行  %d 列"% (ws.title, rows, columns))
        for i in range(1, rows+1):
            sub_list = []
            for j in range(1, columns+1):
                # str = ws.cell(i,j).value
                str = ws.cell(i,j).value if ws.cell(i,j).value is not None else ""
                sub_list.append(str)
            print(sub_list)

# 不要忘记 close
wb.close()
