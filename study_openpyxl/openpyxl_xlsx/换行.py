"""
@author: wanghongliang
@file: 换行.py
@time: 2021/8/17 14:57 
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active # wb.active returns a Worksheet object
ws['A1'] = "Line 1\nLine 2\nLine 3"
ws['A1'].alignment = Alignment(wrapText=True)
wb.save("wrap.xlsx")

