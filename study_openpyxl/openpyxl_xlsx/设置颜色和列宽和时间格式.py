"""
@author: blake.wang
@file: 设置颜色和列宽和时间格式.py
@time: 2021/12/17 15:17 
"""
import os
import time
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def insert_excel(excel_data):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'results'
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        fill = PatternFill("solid", fgColor="E0FFFF")
        excel_title = ["", "keep autolabeling", "changed items", "Accuracy"]
        ws.append(excel_title)
        ws['A1'].fill = fill
        ws['B1'].fill = fill
        ws['C1'].fill = fill
        ws['D1'].fill = fill
        for line in excel_data:
            ws.append(line)
        result_path = os.path.join(os.getcwd(), time.strftime("result_%Y%m%d_%H%M%S.xlsx", time.localtime()))
        wb.save(result_path)
        wb.close()
        print("Result path: ", result_path)
    except Exception as e:
        print("insert_excel error . message: ", e)


# 设置时间格式
def insert_excel_2(sheet1, sheet2_data, close_file_diff, app_close_and_shutdown_diff, emmc_log_close_diff):
    try:
        time_type = datetime.datetime.now().strftime("%Y%m%d_%H%M%S.%f")
        result = "result_" + time_type[0:18] + ".xlsx"
        file_path = os.path.join(os.getcwd(), result)
        wb = Workbook()  # 创建文件对象
        ws = wb.active
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.title = 'Log'
        ws.append(["log time", "utc time", "log path", "log"])
        for line_index, line in enumerate(sheet1):
            for index, item in enumerate(line):
                ws.cell(line_index + 2, index + 1).value = item
                if index == 1:  # 设置时间格式
                    ws.cell(line_index + 2, index + 1).number_format = "mm-dd hh:mm:ss.000"

        # sheet2 数据
        sheet2 = wb.create_sheet("Stat")
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 20
        sheet2.column_dimensions['C'].width = 10
        sheet2.column_dimensions['D'].width = 20
        sheet2.column_dimensions['E'].width = 28
        sheet2.column_dimensions['F'].width = 20
        sheet2.column_dimensions['G'].width = 20

        sheet2.cell(1, 3).value = "Max"
        sheet2.cell(2, 3).value = "Avg"
        sheet2.cell(3, 3).value = "Best"
        sheet2.cell(4, 3).value = "Total"
        sheet2.cell(5, 1).value = "ACC OFF"
        sheet2.cell(5, 2).value = "Shutdown"
        sheet2.cell(5, 4).value = "Close file"
        sheet2.cell(5, 5).value = "App close & shutdown"
        sheet2.cell(5, 6).value = "EMMC log close"
        # sheet2.cell(5, 7).value = "EMMC log close"

        # Close file   Max  Avg  Best  Total
        if len(close_file_diff) > 0:
            sum1 = 0
            for i1 in close_file_diff:
                sum1 = sum1 + i1
            sheet2.cell(1, 4).value = max(close_file_diff)
            sheet2.cell(2, 4).value = sum1 / len(close_file_diff)
            sheet2.cell(3, 4).value = min(close_file_diff)
        sheet2.cell(4, 4).value = len(close_file_diff)

        # App close & shutdown    Max  Avg  Best  Total
        if len(app_close_and_shutdown_diff) > 0:
            sum2 = 0
            for i2 in app_close_and_shutdown_diff:
                sum2 = sum2 + i2
            sheet2.cell(1, 5).value = max(app_close_and_shutdown_diff)
            sheet2.cell(2, 5).value = sum2 / len(app_close_and_shutdown_diff)
            sheet2.cell(3, 5).value = min(app_close_and_shutdown_diff)
        sheet2.cell(4, 5).value = len(app_close_and_shutdown_diff)

        # EMMC log close  Max  Avg  Best  Total
        if len(emmc_log_close_diff) > 0:
            sum4 = 0
            for i4 in emmc_log_close_diff:
                sum4 = sum4 + i4
            sheet2.cell(1, 6).value = max(emmc_log_close_diff)
            sheet2.cell(2, 6).value = sum4 / len(emmc_log_close_diff)
            sheet2.cell(3, 6).value = min(emmc_log_close_diff)
        sheet2.cell(4, 6).value = len(emmc_log_close_diff)

        for line2 in sheet2_data:
            sheet2.append(line2)

        wb.save(file_path)
        wb.close()
        print("result address: ", file_path)
    except Exception as e:
        print("insert_excel error . message: ", e)