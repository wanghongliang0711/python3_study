"""
@author: blake.wang
@file: 插入图片_设置超链接.py
@time: 2022/6/28 13:52 
"""
from openpyxl.drawing.image import Image
from openpyxl import Workbook  # 将数据写入Excel
import os

wb = Workbook()

img_path = r'E:\code\workCode\2022\python-tool\auto-k_app_tool\auto-k_auto_test\auto-k_auto_test_R01\result\1.2.85.4_20220628_132438\211025221631_log\1655762964804_0.3800_0.3300to0.5250_0.3417.jpeg'

ws = wb.active
img = Image(img_path)
new_size = (90, 90)
# 设置图片宽高
img.width, img.height = new_size
ws.add_image(img, 'F1')

# 设置超链接
ws.cell(row=8, column=8, value="666")
ws.cell(row=8, column=8).hyperlink = "./pic/1655762964804_0.3800_0.3300to0.5250_0.3417.jpeg"

wb.save('插入图片.xlsx')
wb.close()


